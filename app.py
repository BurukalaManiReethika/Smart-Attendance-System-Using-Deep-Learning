"""
app.py - Smart Attendance System (Web Version)
------------------------------------------------
Flask backend that:
  - Serves a webcam page (browser camera via getUserMedia)
  - Receives frames from the browser, runs face recognition
  - Marks attendance (once per person per day) into a CSV
  - Lets you enroll new people via browser camera
  - Shows today's attendance list

Run locally:
    python app.py
Deploy on Render: see README_DEPLOY.md
"""

import os
import io
import base64
import pickle
import csv
import json
import smtplib
import threading
from email.mime.text import MIMEText
from datetime import datetime, timedelta

import numpy as np
import cv2
import face_recognition
from flask import Flask, render_template, request, jsonify, send_file, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
KNOWN_FACES_DIR = os.path.join(BASE_DIR, "known_faces")
ATTENDANCE_DIR = os.path.join(BASE_DIR, "attendance_records")
ENCODINGS_FILE = os.path.join(BASE_DIR, "encodings.pickle")
EMAILS_FILE = os.path.join(BASE_DIR, "emails.json")

# Anyone marked present after this time (24h "HH:MM") is tagged "Late" instead of "Present".
# Override with the LATE_AFTER environment variable, e.g. LATE_AFTER=09:15
LATE_AFTER = os.environ.get("LATE_AFTER", "09:30")

TOLERANCE = 0.5
MODEL = "hog"

# --- Email (Gmail SMTP) config, read from environment variables ---
SMTP_EMAIL = os.environ.get("SMTP_EMAIL")        # your Gmail address, e.g. yourname@gmail.com
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")  # Gmail App Password (NOT your normal password)
NOTIFY_ADMIN_EMAIL = os.environ.get("NOTIFY_ADMIN_EMAIL")  # optional: also notify this address on every mark
EMAIL_ENABLED = bool(SMTP_EMAIL and SMTP_PASSWORD)

os.makedirs(KNOWN_FACES_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

# In-memory cache of encodings, loaded at startup / after enrollment
known_encodings = []
known_names = []


def load_encodings():
    global known_encodings, known_names
    if os.path.exists(ENCODINGS_FILE):
        with open(ENCODINGS_FILE, "rb") as f:
            data = pickle.load(f)
        known_encodings = data.get("encodings", [])
        known_names = data.get("names", [])
    else:
        known_encodings, known_names = [], []


def save_encodings():
    with open(ENCODINGS_FILE, "wb") as f:
        pickle.dump({"encodings": known_encodings, "names": known_names}, f)


def load_emails():
    if os.path.exists(EMAILS_FILE):
        with open(EMAILS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_email_for_person(name, email):
    emails = load_emails()
    if email:
        emails[name] = email
        with open(EMAILS_FILE, "w") as f:
            json.dump(emails, f, indent=2)


def send_email_async(to_address, subject, body):
    """Send an email in a background thread so it never blocks/slow the request."""
    if not EMAIL_ENABLED or not to_address:
        return

    def _send():
        try:
            msg = MIMEText(body)
            msg["Subject"] = subject
            msg["From"] = SMTP_EMAIL
            msg["To"] = to_address
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
                server.login(SMTP_EMAIL, SMTP_PASSWORD)
                server.sendmail(SMTP_EMAIL, [to_address], msg.as_string())
            print(f"[EMAIL SENT] Successfully sent to {to_address}")
        except Exception as e:
            print(f"[EMAIL ERROR] Could not send to {to_address}: {type(e).__name__}: {e}")

    threading.Thread(target=_send, daemon=True).start()


def notify_attendance_marked(name):
    """Send confirmation email to the person, and optionally to an admin address."""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    emails = load_emails()
    person_email = emails.get(name)

    print(f"[EMAIL DEBUG] notify_attendance_marked called for '{name}'. "
          f"EMAIL_ENABLED={EMAIL_ENABLED}, saved_emails={emails}, person_email={person_email}")

    if not EMAIL_ENABLED:
        print("[EMAIL SKIP] SMTP_EMAIL / SMTP_PASSWORD not set in environment.")
        return

    if person_email:
        send_email_async(
            person_email,
            "Attendance Marked — Smart Attendance System",
            f"Hi {name},\n\nYour attendance was marked present on {now_str}.\n\n"
            f"— Smart Attendance System"
        )
    else:
        print(f"[EMAIL SKIP] No email saved for '{name}'. Available names: {list(emails.keys())}")

    if NOTIFY_ADMIN_EMAIL:
        send_email_async(
            NOTIFY_ADMIN_EMAIL,
            f"Attendance: {name} marked present",
            f"{name} was marked present at {now_str}."
        )


def get_all_attendance_records():
    """Read every attendance CSV file and return a combined text summary."""
    all_records = []
    for date_str in sorted(list_attendance_dates()):
        for row in read_attendance_csv(date_str):
            name, date, time, status = row
            tag = " (Late)" if status == "Late" else ""
            all_records.append(f"{name} — {date} at {time}{tag}")
    return all_records


def ask_chatbot(question):
    """
    Free, rule-based attendance assistant — no external API, no cost.
    Understands common attendance questions using keyword matching.
    """
    q = question.lower().strip()

    enrolled_people = sorted(set(known_names))
    today_marked = sorted(load_already_marked())
    records = get_all_attendance_records()
    today_str = datetime.now().strftime("%Y-%m-%d")

    # --- Who is present today ---
    if any(phrase in q for phrase in ["who is present", "who's present", "present today", "who all present", "who came today"]):
        if today_marked:
            return f"Marked present today ({today_str}):\n" + "\n".join(f"• {n}" for n in today_marked)
        return f"No one has been marked present yet today ({today_str})."

    # --- How many present today / count ---
    if any(phrase in q for phrase in ["how many", "count", "total present", "number of people"]):
        return f"{len(today_marked)} out of {len(enrolled_people)} enrolled people are marked present today ({today_str})."

    # --- Who is enrolled ---
    if any(phrase in q for phrase in ["who is enrolled", "who's enrolled", "list of people", "who are enrolled", "registered people"]):
        if enrolled_people:
            return "Enrolled people in the system:\n" + "\n".join(f"• {n}" for n in enrolled_people)
        return "No one is enrolled in the system yet."

    # --- Was <name> marked / did <name> attend ---
    for name in enrolled_people:
        if name.lower().replace("_", " ") in q or name.lower() in q:
            if name in today_marked:
                # find their exact time from today's records
                time_str = ""
                for r in records:
                    if r.startswith(name) and today_str in r:
                        time_str = r.split(" at ")[-1] if " at " in r else ""
                return f"Yes — {name} was marked present today{f' at {time_str}' if time_str else ''}."
            else:
                # check full history
                past = [r for r in records if r.startswith(name)]
                if past:
                    return f"{name} was NOT marked present today. Most recent record: {past[-1]}"
                return f"{name} is enrolled but has no attendance records yet."

    # --- Attendance history / all records ---
    if any(phrase in q for phrase in ["history", "all records", "full attendance", "every record"]):
        if records:
            return "Full attendance history:\n" + "\n".join(f"• {r}" for r in records[-20:])
        return "No attendance records exist yet."

    # --- Fallback: didn't understand ---
    return (
        "I can answer questions like:\n"
        "• Who is present today?\n"
        "• How many people are present?\n"
        "• Who is enrolled in the system?\n"
        "• Was <name> marked today?\n\n"
        "Try rephrasing your question using one of these patterns."
    )


def rebuild_encodings_from_disk():
    """Re-scan known_faces/ and rebuild encodings.pickle from scratch."""
    global known_encodings, known_names
    new_encodings, new_names = [], []

    for person_name in os.listdir(KNOWN_FACES_DIR):
        person_dir = os.path.join(KNOWN_FACES_DIR, person_name)
        if not os.path.isdir(person_dir):
            continue
        for image_file in os.listdir(person_dir):
            if not image_file.lower().endswith((".jpg", ".jpeg", ".png")):
                continue
            image_path = os.path.join(person_dir, image_file)
            image = face_recognition.load_image_file(image_path)
            boxes = face_recognition.face_locations(image, model=MODEL)
            encs = face_recognition.face_encodings(image, boxes)
            if encs:
                new_encodings.append(encs[0])
                new_names.append(person_name)

    known_encodings, known_names = new_encodings, new_names
    save_encodings()


def decode_base64_image(data_url):
    """Convert a 'data:image/jpeg;base64,...' string into an OpenCV BGR image."""
    header, encoded = data_url.split(",", 1)
    img_bytes = base64.b64decode(encoded)
    np_arr = np.frombuffer(img_bytes, dtype=np.uint8)
    frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    return frame


def get_csv_path_for_date(date_str):
    return os.path.join(ATTENDANCE_DIR, f"attendance_{date_str}.csv")


def get_today_csv_path():
    today = datetime.now().strftime("%Y-%m-%d")
    return get_csv_path_for_date(today)


def compute_status(dt):
    """Return 'Present' or 'Late' based on LATE_AFTER cutoff (HH:MM, 24h)."""
    try:
        cutoff = datetime.strptime(LATE_AFTER, "%H:%M").time()
    except ValueError:
        cutoff = datetime.strptime("09:30", "%H:%M").time()
    return "Late" if dt.time() > cutoff else "Present"


def read_attendance_csv(date_str):
    """Read one day's CSV. Returns list of [name, date, time, status].
    Older files saved before the Status column existed are filled in as 'Present'."""
    csv_path = get_csv_path_for_date(date_str)
    rows = []
    if os.path.exists(csv_path):
        with open(csv_path, "r", newline="") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if not row:
                    continue
                if len(row) >= 4:
                    rows.append([row[0], row[1], row[2], row[3]])
                elif len(row) == 3:
                    rows.append([row[0], row[1], row[2], "Present"])
    return rows


def list_attendance_dates():
    """All dates (YYYY-MM-DD) that have an attendance file, most recent first."""
    dates = []
    if os.path.isdir(ATTENDANCE_DIR):
        for filename in os.listdir(ATTENDANCE_DIR):
            if filename.startswith("attendance_") and filename.endswith(".csv"):
                dates.append(filename[len("attendance_"):-len(".csv")])
    return sorted(dates, reverse=True)


def load_already_marked():
    marked = set()
    for row in read_attendance_csv(datetime.now().strftime("%Y-%m-%d")):
        marked.add(row[0])
    return marked


def mark_attendance(name):
    csv_path = get_today_csv_path()
    marked = load_already_marked()
    if name in marked:
        return False
    file_exists = os.path.exists(csv_path)
    now = datetime.now()
    status = compute_status(now)
    with open(csv_path, "a", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["Name", "Date", "Time", "Status"])
        writer.writerow([name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), status])
    return True


def build_dashboard_stats(days=14):
    """Daily present-counts for the last N days, plus per-person attendance rate
    over the days the system actually recorded attendance."""
    all_dates = list_attendance_dates()
    recent_dates = sorted(all_dates)[-days:] if all_dates else []

    daily_counts = []
    for d in recent_dates:
        rows = read_attendance_csv(d)
        daily_counts.append({
            "date": d,
            "present": sum(1 for r in rows if r[3] == "Present"),
            "late": sum(1 for r in rows if r[3] == "Late"),
        })

    enrolled = sorted(set(known_names))
    tracked_dates = all_dates  # every day with a record, for the rate calculation
    per_person = []
    for person in enrolled:
        days_present = 0
        for d in tracked_dates:
            rows = read_attendance_csv(d)
            if any(r[0] == person for r in rows):
                days_present += 1
        total_days = len(tracked_dates) if tracked_dates else 1
        rate = round((days_present / total_days) * 100) if tracked_dates else 0
        per_person.append({"name": person, "days_present": days_present, "total_days": len(tracked_dates), "rate": rate})

    per_person.sort(key=lambda p: p["rate"], reverse=True)

    return {
        "daily_counts": daily_counts,
        "per_person": per_person,
        "total_enrolled": len(enrolled),
        "total_tracked_days": len(tracked_dates),
    }


# ---------------------------------------------------------------- Routes

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/enroll")
def enroll_page():
    return render_template("enroll.html")


@app.route("/attendance")
def attendance_page():
    requested_date = request.args.get("date", "").strip()
    today_str = datetime.now().strftime("%Y-%m-%d")
    date_str = requested_date if requested_date else today_str

    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        date_str = today_str

    rows = read_attendance_csv(date_str)
    available_dates = list_attendance_dates()

    # prev/next available date for quick navigation
    prev_date = next((d for d in available_dates if d < date_str), None)
    next_date = next((d for d in reversed(available_dates) if d > date_str), None)

    return render_template(
        "attendance.html",
        rows=rows,
        selected_date=date_str,
        today=today_str,
        is_today=(date_str == today_str),
        available_dates=available_dates,
        prev_date=prev_date,
        next_date=next_date,
        late_after=LATE_AFTER,
    )


@app.route("/dashboard")
def dashboard_page():
    return render_template("dashboard.html", late_after=LATE_AFTER)


@app.route("/api/dashboard-data")
def api_dashboard_data():
    days = request.args.get("days", 14, type=int)
    return jsonify(build_dashboard_stats(days=days))


@app.route("/api/export")
def api_export():
    """Download attendance as CSV or Excel — a single date, or the full history."""
    fmt = request.args.get("format", "csv").lower()
    scope = request.args.get("scope", "day")  # "day" or "all"
    date_str = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))

    if scope == "all":
        rows = []
        for d in sorted(list_attendance_dates()):
            rows.extend(read_attendance_csv(d))
        filename_base = "attendance_full_history"
    else:
        rows = read_attendance_csv(date_str)
        filename_base = f"attendance_{date_str}"

    header = ["Name", "Date", "Time", "Status"]

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F7A5C", end_color="2F7A5C", fill_type="solid")
        for row in rows:
            ws.append(row)
        for col_cells in ws.columns:
            width = max(len(str(c.value)) for c in col_cells if c.value is not None) + 4
            ws.column_dimensions[col_cells[0].column_letter].width = max(width, 12)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{filename_base}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"},
    )


@app.route("/api/recognize", methods=["POST"])
def api_recognize():
    """Receive one frame from the browser, recognize faces, mark attendance."""
    payload = request.get_json(force=True)
    image_data = payload.get("image")
    if not image_data:
        return jsonify({"error": "No image provided"}), 400

    frame = decode_base64_image(image_data)
    if frame is None:
        return jsonify({"error": "Could not decode image"}), 400

    small = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
    rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_small, model=MODEL)
    face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

    results = []
    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        name = "Unknown"
        marked_now = False

        if known_encodings:
            distances = face_recognition.face_distance(known_encodings, face_encoding)
            best_idx = int(distances.argmin())
            if distances[best_idx] <= TOLERANCE:
                name = known_names[best_idx]
                marked_now = mark_attendance(name)
                if marked_now:
                    notify_attendance_marked(name)

        # scale box coords back up (we resized by 0.5)
        results.append({
            "name": name,
            "box": [int(top * 2), int(right * 2), int(bottom * 2), int(left * 2)],
            "marked_now": marked_now
        })

    return jsonify({"faces": results, "marked_today": sorted(load_already_marked())})


@app.route("/api/enroll", methods=["POST"])
def api_enroll():
    """Receive a name + list of base64 images, save them, rebuild encodings."""
    payload = request.get_json(force=True)
    name = payload.get("name", "").strip().replace(" ", "_")
    email = payload.get("email", "").strip()
    images = payload.get("images", [])

    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not images:
        return jsonify({"error": "No images provided"}), 400

    person_dir = os.path.join(KNOWN_FACES_DIR, name)
    os.makedirs(person_dir, exist_ok=True)

    saved = 0
    for i, img_data in enumerate(images):
        frame = decode_base64_image(img_data)
        if frame is None:
            continue
        # quick check: does this photo actually contain a face?
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        boxes = face_recognition.face_locations(rgb, model=MODEL)
        if not boxes:
            continue
        filename = os.path.join(person_dir, f"{name}_{i}.jpg")
        cv2.imwrite(filename, frame)
        saved += 1

    if saved == 0:
        return jsonify({"error": "No faces detected in captured photos. Try again with better lighting."}), 400

    save_email_for_person(name, email)
    rebuild_encodings_from_disk()

    return jsonify({"message": f"Enrolled '{name}' with {saved} photo(s).", "total_people": len(set(known_names))})


@app.route("/chat")
def chat_page():
    return render_template("chat.html")


@app.route("/api/chat", methods=["POST"])
def api_chat():
    payload = request.get_json(force=True)
    question = payload.get("question", "").strip()
    if not question:
        return jsonify({"error": "Question is required"}), 400

    answer = ask_chatbot(question)
    return jsonify({"answer": answer})


@app.route("/api/people", methods=["GET"])
def api_people():
    return jsonify({"people": sorted(set(known_names))})


load_encodings()
print(f"[STARTUP] EMAIL_ENABLED={EMAIL_ENABLED} | SMTP_EMAIL={'SET' if SMTP_EMAIL else 'NOT SET'} | "
      f"SMTP_PASSWORD={'SET' if SMTP_PASSWORD else 'NOT SET'} | "
      f"known_people={sorted(set(known_names))}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
